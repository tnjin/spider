import urllib.request
from bs4 import BeautifulSoup
import http.cookiejar
from openpyxl import Workbook
from openpyxl import load_workbook
import os

#抓取汽车之家的车型数据

def load_html(openner, url):
    res = openner.open(url)
    page = res.read()
    html_str = page.decode("gbk")
    res.close()
    # print(html)
    return html_str


def decode_car(html):
    soup = BeautifulSoup(html, "html.parser")
    # s3895
    car_info = soup.select("div.list-cont-main")
    cars = []
    for info in car_info:

        key_ele = "电 动 机："
        key_range = "续航里程："
        a = info.select("div.main-title > a")
        if len(a) == 1:
            name = a[0].text
        else:
            print("no name found ")
        score = info.select("span.score-number")
        if len(score) == 1:
            score_num = score[0].text
        else:
            score_num = "NO"
        car = dict()
        car["name"] = name
        car["score"] = score_num
        ul = info.select("ul")
        for li in ul:
            des = li.text
            # print(des)
            if des.find(key_ele) > -1:
                decode_ele_car(des, car)
            else:
                decode_oil_car(des,car)
            cars.append(car)
    return cars


def decode_ele_car(des, car):
    #print(des)
    key_range = "续航里程："
    key_eleng = "电 动 机："
    key_charge = "充电时间："
    key_color = "外观颜色："

    car["engine"] = "电动"
    car["struct"] = ""
    car["gear"] = ""
    car["level"] = ""
    start = des.find(key_range)
    end = des.find(key_eleng)
    car["range"] = des[start+ len(key_range):end]
    start = end
    end = des.find(key_charge)
    car["horse_power"] = des[start+len(key_eleng):end]
    start = end
    end = des.find(key_color)
    car["charge_time"] = des[start+len(key_charge):end]
    car["color"] = des[end + len(key_color):]
    return car


def decode_oil_car(des, car):
    key_level = "级  别："
    key_struct = "车身结构："
    key_eng = "发 动 机："
    key_gear = "变 速 箱："
    key_color = "外观颜色："
    key_color_splitor = "色"
    car["engine"] = "内燃机"
    car["range"] = ""
    car["horse_power"] = ""
    car["charge_time"] = ""
    start = (des.find(key_level) + len(key_level))
    end = des.find(key_struct)
    level = des[start:end]
    # print(level)
    car["level"] = level
    start = end + len(key_struct)
    end = des.find(key_eng)
    struct = des[start:end]
    # print(struct)
    car["struct"] = struct
    start = end + len(key_eng)
    end = des.find(key_gear)
    eng = des[start: end]
    # print(eng)
    car["engine"] = eng
    start = end + len(key_gear)
    end = des.find(key_color)
    gear = des[start: end]
    # print(gear)
    car["gear"] = gear
    start = end + len(key_color)
    colors = des[start:-1]
    # print(colors)
    car["color"] = colors.split(key_color_splitor)
    return car


def decode_car_list(html):
    soup = BeautifulSoup(html, "html.parser")
    letter_item = soup.select("div.cartree-letter")
    ul_item = soup.select("ul")
    # for item in letter_item:
        # print(item.text)
    print(len(ul_item))
    cars = []
    for ul in ul_item:
        for li in ul:
            car = dict()
            car["brand"] = li.text
            car["url"] = li.a["href"]
            # print(car)
            cars.append(car)
    # print(len(cars))
    return cars


def write_file(cars):
    if cars:
        with open("C:\\Users\\Tjz\\Desktop\\开发\\car.txt", "a+") as file:
            for car in cars:
                file.write(str(car))
                file.write("\n")


def writ_xls(cars ,path):
    if os.path.exists(path):
        wb = load_workbook(path)
    else:
        wb = Workbook()
        wb.create_sheet("车型")
        wb.save(path)
    sheet = wb.get_sheet_by_name("车型")
    if sheet:
        print("直接插入")
    else:
        sheet = wb.create_sheet("车型")
    for car in cars:
        sheet.append(car_list(car))
    wb.save(path)


def car_list(car):
    row = [car["name"],
           car["level"],
           car["struct"],
           car["engine"],
           car["gear"],
           car["range"],
           car["score"]]
    return row

if __name__ == '__main__':
        url = r"http://car.autohome.com.cn/AsLeftMenu/As_LeftListNew.ashx?typeId=1%20&brandId=0%20&fctId=0%20&seriesId=0"
        # print(url)
        cj = http.cookiejar.CookieJar()
        openner = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))
        _str = load_html(openner,url)
        _str = _str[len('document.writeln("'):-3]
        #print(_str)
        car_pages = decode_car_list(_str)
        base_url = r"http://car.autohome.com.cn/"
        i = 1
        total = 0
        count = len(car_pages)
        for page in car_pages:
            url = base_url + page["url"]
            _str = load_html(openner, url)
            # print(_str)
            _cars = decode_car(_str)
            # write_file(_cars)
            writ_xls(_cars,"C:\\Users\\Tjz\\Desktop\\开发\\car.xlsx")
            total += len(_cars)
            print("total:%d|cur:%d|cars:%d" % (count, i, total))
            i += 1


