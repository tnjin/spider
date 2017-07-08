import urllib.request
from bs4 import BeautifulSoup
import http.cookiejar
import chardet
import gzip
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import re



def load_html(openner, url):
    res = openner.open(url)
    page = res.read()
    encoding = chardet.detect(page)
    print(encoding)
    html_str = page.decode("gbk")
    res.close()
    # print(html)
    return html_str

def load_html2(openner, url):
    res = openner.open(url)
    page = gzip.decompress(res.read())
    #encoding = chardet.detect(page)
    #print(encoding)
    html_str = page.decode("gb18030")
    res.close()
    # print(html)
    return html_str


def get_default_post():
    post = dict()
    post["title"] = "null-t"
    post["author"] = "null-t"
    post["time"] = "null-t"
    post["url"] = "null-t"
    post["content"] = "null-t"
    post["reply"] = -1
    post["click"] = -1
    post["area"] = "null-t"
    return post


def decode_first_html(html):
    soup = BeautifulSoup(html, "html.parser")
    titles = soup.select("dl.list_dl")
    posts = []
    for title in titles:
        try:
            post = get_default_post()
            text_title = title.select("a.a_topic")
            text_author = title.select("a.linkblack")
            text_post_time = title.select("span.tdate")
            print_data(text_title)
            post["title"] = text_title[0].text
            post["url"] = text_title[0].get("href")
            post["author"] = text_author[0].text
            post["time"] = text_post_time[0].text
            posts.append(post)
            #print(text_title[0].get("href"))
            #print_data(text_author)
            #print_data(text_post_time)
        except Exception :
            print(post)

    return posts

def decode_sec_html(html,post):
    #print(html)
    soup = BeautifulSoup(html, "html.parser")
    clickNum = soup.select("#x-views")[0].text
    repNum = soup.select("#x-replys")[0].text
    post["click"] = clickNum
    post["reply"] = repNum
    links = soup.select("li > a.c01439a")
    #print(links)
    for link in links:
        if(link.get("title") and "查看该地区论坛" == link.get("title") ):
           # post["area"] = link.text
           #print(link.get("title"))
           #print(link.text)
           post["area"] = link.text
           break
    content = soup.select("div.tz-paragraph")
    if(len(content)>0):
        c = content[0].text
        patter = re.compile(r"<span class.*\"></span>")
        c = re.sub(pattern=patter,repl="", string=c)
        post["content"] = c

def writ_xls(cars ,path):
    if os.path.exists(path):
        wb = load_workbook(path)
    else:
        wb = Workbook()
        wb.create_sheet("帖子")
        wb.save(path)
    sheet = wb.get_sheet_by_name("帖子")
    if sheet:
        print("直接插入")
    else:
        sheet = wb.create_sheet("帖子")
    for car in cars:
        sheet.append(car_list(car))
    wb.save(path)


def write_file(cars):
    if cars:
        with open("C:\\Users\\Tjz\\Desktop\\开发\\car2.txt", mode="a+", encoding="gb18030") as file:
            for car in cars:
                file.write(psotStr(car))
                file.write("\n")


def car_list(car):
    row = [car["title"],
           car["author"],
           car["time"],
           car["click"],
           car["reply"],
           car["area"],
           car["content"],
           car["url"]]
    return row


def psotStr(post):
    s = post["title"]+"|"+post["author"]+"|"+post["time"]+"|"+ post["click"]+"|"+post["reply"]+"|"+post["area"]+"|"+post["content"]+"|"+post["url"]
    return s

def print_data(data):
    for d in data :
        print(d.text)


def poll(baseUrl):
    cj = http.cookiejar.CookieJar()
    openner = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))
    openner.addheaders = [('User-agent', 'Mozilla/5.0'), ('Accept-Encoding', 'gzip, deflate')]
    _str = load_html2(openner, baseUrl)
    _posts = decode_first_html(_str)
    for post in _posts:
        secUrl = base_url + post["url"]
        post["url"] = secUrl
        print(secUrl)
        _str = load_html2(openner, secUrl)
        decode_sec_html(_str, post)
    #write_file(_posts)
    writ_xls(_posts, "C:\\Users\\Tjz\\Desktop\\开发\\car2.xlsx")

if __name__ == '__main__':
        base_url = r"http://club.autohome.com.cn"
        url = r"http://club.autohome.com.cn/bbs/forum-c-692-%d.html"
        # print(url)
        i = 0
        while i < 50:
            i = i + 1
            temp = url % i
            print(temp)
            poll(temp)

